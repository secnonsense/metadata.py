import tkinter as tk
from tkinter import filedialog, scrolledtext, ttk, messagebox
import os
import mimetypes
from PIL import Image, ExifTags
from mutagen.mp3 import MP3
from mutagen.flac import FLAC
from mutagen.oggvorbis import OggVorbis
from mutagen.mp4 import MP4
from mutagen.asf import ASF
from mutagen.wavpack import WavPack
from mutagen.aiff import AIFF
from mutagen.trueaudio import TrueAudio
from mutagen.monkeysaudio import MonkeysAudio
from mutagen.musepack import Musepack
from mutagen.optimfrog import OptimFROG
from mutagen.aac import AAC
from mutagen.oggopus import OggOpus
from mutagen.id3 import ID3NoHeaderError
import PyPDF2
import docx # For .docx files
import openpyxl # For .xlsx files
from pptx import Presentation # For .pptx files

class FileMetadataExtractor:
    def __init__(self, master):
        self.master = master
        master.title("File Metadata Extractor")
        master.geometry("800x600")
        master.resizable(True, True)

        # Configure grid for responsiveness
        master.grid_rowconfigure(2, weight=1)
        master.grid_columnconfigure(1, weight=1)

        # --- File Selection Frame ---
        file_frame = ttk.LabelFrame(master, text="File Selection")
        file_frame.grid(row=0, column=0, columnspan=2, padx=10, pady=10, sticky="ew")
        file_frame.grid_columnconfigure(1, weight=1)

        self.file_path = tk.StringVar()
        self.file_path.set("No file selected")

        ttk.Label(file_frame, text="File Path:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.path_entry = ttk.Entry(file_frame, textvariable=self.file_path, width=70, state='readonly')
        self.path_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        ttk.Button(file_frame, text="Browse", command=self.browse_file).grid(row=0, column=2, padx=5, pady=5)

        # --- File Type Selection Frame ---
        type_frame = ttk.LabelFrame(master, text="File Type")
        type_frame.grid(row=1, column=0, columnspan=2, padx=10, pady=5, sticky="ew")
        type_frame.grid_columnconfigure(1, weight=1)

        ttk.Label(type_frame, text="Detected Type:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.detected_type = tk.StringVar()
        self.detected_type.set("N/A")
        ttk.Label(type_frame, textvariable=self.detected_type).grid(row=0, column=1, padx=5, pady=5, sticky="w")

        ttk.Label(type_frame, text="Manual Override:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.manual_type_combo = ttk.Combobox(type_frame, state="readonly", width=30)
        self.manual_type_combo['values'] = [
            "Auto-Detect", "Image", "Audio/Video", "PDF",
            "Word Document (DOCX)", "Excel Spreadsheet (XLSX)", "PowerPoint Presentation (PPTX)",
            "Generic"
        ]
        self.manual_type_combo.set("Auto-Detect")
        self.manual_type_combo.grid(row=1, column=1, padx=5, pady=5, sticky="ew")

        ttk.Button(type_frame, text="Extract Metadata", command=self.extract_metadata).grid(row=1, column=2, padx=5, pady=5)

        # --- Metadata Display Area ---
        self.metadata_display = scrolledtext.ScrolledText(master, wrap=tk.WORD, width=90, height=20, font=("Courier New", 10))
        self.metadata_display.grid(row=2, column=0, columnspan=2, padx=10, pady=10, sticky="nsew")
        self.metadata_display.insert(tk.END, "Select a file and click 'Extract Metadata' to view details.")
        self.metadata_display.config(state=tk.DISABLED) # Make it read-only

    def browse_file(self):
        """Opens a file dialog for the user to select a file."""
        filepath = filedialog.askopenfilename()
        if filepath:
            self.file_path.set(filepath)
            self.detected_type.set(self._detect_mime_type(filepath))
            self.metadata_display.config(state=tk.NORMAL)
            self.metadata_display.delete(1.0, tk.END)
            self.metadata_display.insert(tk.END, f"File selected: {filepath}\nDetected type: {self.detected_type.get()}\n\nClick 'Extract Metadata' to proceed.")
            self.metadata_display.config(state=tk.DISABLED)

    def _detect_mime_type(self, filepath):
        """Detects the MIME type of a file."""
        mime_type, _ = mimetypes.guess_type(filepath)
        if mime_type:
            return mime_type
        # Add some common office document extensions if mimetypes doesn't catch them
        ext = os.path.splitext(filepath)[1].lower()
        if ext == ".docx":
            return "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        elif ext == ".xlsx":
            return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        elif ext == ".pptx":
            return "application/vnd.openxmlformats-officedocument.presentationml.presentation"
        return "application/octet-stream" # Default for unknown

    def extract_metadata(self):
        """Extracts and displays metadata based on file type."""
        filepath = self.file_path.get()
        if not os.path.exists(filepath):
            messagebox.showerror("Error", "Please select a valid file first.")
            return

        self.metadata_display.config(state=tk.NORMAL)
        self.metadata_display.delete(1.0, tk.END) # Clear previous content

        metadata_dict = {}

        # Get general file system metadata
        metadata_dict.update(self._get_general_metadata(filepath))

        # Determine extraction method
        manual_selection = self.manual_type_combo.get()
        detected_mime = self.detected_type.get()
        file_type_category = "Generic"

        # Logic to determine file type category based on manual selection or detected MIME type
        if manual_selection == "Image" or (manual_selection == "Auto-Detect" and detected_mime.startswith("image/")):
            file_type_category = "Image"
        elif manual_selection == "Audio/Video" or (manual_selection == "Auto-Detect" and (detected_mime.startswith("audio/") or detected_mime.startswith("video/"))):
            file_type_category = "Audio/Video"
        elif manual_selection == "PDF" or (manual_selection == "Auto-Detect" and detected_mime == "application/pdf"):
            file_type_category = "PDF"
        elif manual_selection == "Word Document (DOCX)" or (manual_selection == "Auto-Detect" and detected_mime == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"):
            file_type_category = "Word Document"
        elif manual_selection == "Excel Spreadsheet (XLSX)" or (manual_selection == "Auto-Detect" and detected_mime == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"):
            file_type_category = "Excel Spreadsheet"
        elif manual_selection == "PowerPoint Presentation (PPTX)" or (manual_selection == "Auto-Detect" and detected_mime == "application/vnd.openxmlformats-officedocument.presentationml.presentation"):
            file_type_category = "PowerPoint Presentation"
        # Else, it remains "Generic"

        self.metadata_display.insert(tk.END, f"--- Metadata for: {os.path.basename(filepath)} ---\n")
        self.metadata_display.insert(tk.END, f"--- Category: {file_type_category} ---\n\n")

        # Extract specific metadata based on determined category
        if file_type_category == "Image":
            metadata_dict.update(self._get_image_metadata(filepath))
        elif file_type_category == "Audio/Video":
            metadata_dict.update(self._get_audio_video_metadata(filepath))
        elif file_type_category == "PDF":
            metadata_dict.update(self._get_pdf_metadata(filepath))
        elif file_type_category == "Word Document":
            metadata_dict.update(self._get_docx_metadata(filepath))
        elif file_type_category == "Excel Spreadsheet":
            metadata_dict.update(self._get_xlsx_metadata(filepath))
        elif file_type_category == "PowerPoint Presentation":
            metadata_dict.update(self._get_pptx_metadata(filepath))
        # For "Generic", only general metadata is displayed

        self._display_metadata(metadata_dict)
        self.metadata_display.config(state=tk.DISABLED)

    def _get_general_metadata(self, filepath):
        """Extracts general file system metadata."""
        general_meta = {}
        try:
            stat_info = os.stat(filepath)
            general_meta["File Size"] = f"{stat_info.st_size / (1024 * 1024):.2f} MB ({stat_info.st_size} bytes)"
            general_meta["Creation Time"] = f"{os.path.ctime(stat_info.st_ctime)}"
            general_meta["Modification Time"] = f"{os.path.ctime(stat_info.st_mtime)}"
            general_meta["Last Access Time"] = f"{os.path.ctime(stat_info.st_atime)}"
            general_meta["Owner UID"] = stat_info.st_uid
            general_meta["Group GID"] = stat_info.st_gid
            general_meta["Permissions"] = oct(stat_info.st_mode & 0o777)
        except Exception as e:
            general_meta["Error (General Metadata)"] = str(e)
        return general_meta

    def _get_image_metadata(self, filepath):
        """Extracts EXIF metadata from images using Pillow."""
        image_meta = {}
        try:
            with Image.open(filepath) as img:
                image_meta["Image Format"] = img.format
                image_meta["Image Mode"] = img.mode
                image_meta["Image Size"] = f"{img.width}x{img.height} pixels"

                if img.info:
                    for key, value in img.info.items():
                        if key not in ["exif", "icc_profile"]: # Avoid raw binary data
                            image_meta[f"Image Info: {key}"] = value

                exif_data = img._getexif()
                if exif_data:
                    for tag, value in exif_data.items():
                        decoded = ExifTags.TAGS.get(tag, tag)
                        if isinstance(value, bytes):
                            try:
                                value = value.decode('utf-8', errors='ignore')
                            except:
                                pass # Keep as bytes if decoding fails
                        image_meta[f"EXIF: {decoded}"] = value
        except Exception as e:
            image_meta["Error (Image Metadata)"] = str(e)
        return image_meta

    def _get_audio_video_metadata(self, filepath):
        """Extracts audio/video metadata using Mutagen."""
        audio_video_meta = {}
        try:
            # Try to load with various Mutagen handlers
            audio = None
            try:
                # Specific handlers for common audio/video types
                if filepath.lower().endswith((".mp3")):
                    audio = MP3(filepath)
                elif filepath.lower().endswith((".flac")):
                    audio = FLAC(filepath)
                elif filepath.lower().endswith((".ogg", ".oga")):
                    audio = OggVorbis(filepath)
                elif filepath.lower().endswith((".m4a", ".mp4", ".m4v")):
                    audio = MP4(filepath)
                elif filepath.lower().endswith((".wma", ".wmv", ".asf")):
                    audio = ASF(filepath)
                elif filepath.lower().endswith((".wv")):
                    audio = WavPack(filepath)
                elif filepath.lower().endswith((".aiff", ".aif")):
                    audio = AIFF(filepath)
                elif filepath.lower().endswith((".tta")):
                    audio = TrueAudio(filepath)
                elif filepath.lower().endswith((".ape")):
                    audio = MonkeysAudio(filepath)
                elif filepath.lower().endswith((".mpc")):
                    audio = Musepack(filepath)
                elif filepath.lower().endswith((".ofr")):
                    audio = OptimFROG(filepath)
                elif filepath.lower().endswith((".aac")):
                    audio = AAC(filepath)
                elif filepath.lower().endswith((".opus")):
                    audio = OggOpus(filepath)
                else:
                    # Fallback for other types that mutagen might handle generically
                    from mutagen import File
                    audio = File(filepath) # This will try to auto-detect
            except ID3NoHeaderError:
                audio_video_meta["Info"] = "No ID3 header found, trying generic audio/video parsing."
                try:
                    from mutagen import File
                    audio = File(filepath)
                except Exception:
                    audio = None # Still no luck
            except Exception as e:
                audio_video_meta["Info"] = f"Specific Mutagen handler failed: {e}. Trying generic."
                try:
                    from mutagen import File
                    audio = File(filepath)
                except Exception:
                    audio = None

            if audio:
                audio_video_meta["Length (seconds)"] = f"{audio.info.length:.2f}" if hasattr(audio.info, 'length') else "N/A"
                audio_video_meta["Bitrate (kbps)"] = f"{audio.info.bitrate / 1000:.0f}" if hasattr(audio.info, 'bitrate') else "N/A"
                audio_video_meta["Channels"] = audio.info.channels if hasattr(audio.info, 'channels') else "N/A"
                audio_video_meta["Sample Rate (Hz)"] = audio.info.sample_rate if hasattr(audio.info, 'sample_rate') else "N/A"

                for key, value in audio.items():
                    # Mutagen tags can be complex objects, convert to string for display
                    audio_video_meta[f"Tag: {key}"] = str(value)
            else:
                audio_video_meta["Info"] = "Could not parse audio/video metadata with Mutagen."

        except Exception as e:
            audio_video_meta["Error (Audio/Video Metadata)"] = str(e)
        return audio_video_meta

    def _get_pdf_metadata(self, filepath):
        """Extracts metadata from PDF files using PyPDF2."""
        pdf_meta = {}
        try:
            with open(filepath, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                info = reader.metadata
                if info:
                    for key, value in info.items():
                        # PyPDF2 metadata keys start with '/'
                        pdf_meta[f"PDF: {key.replace('/', '')}"] = value
                pdf_meta["PDF: Number of Pages"] = len(reader.pages)
        except Exception as e:
            pdf_meta["Error (PDF Metadata)"] = str(e)
        return pdf_meta

    def _get_docx_metadata(self, filepath):
        """Extracts metadata from DOCX files using python-docx."""
        docx_meta = {}
        try:
            document = docx.Document(filepath)
            core_properties = document.core_properties
            
            # Extract common core properties
            docx_meta["DOCX: Title"] = core_properties.title
            docx_meta["DOCX: Author"] = core_properties.author
            docx_meta["DOCX: Last Modified By"] = core_properties.last_modified_by
            docx_meta["DOCX: Revision"] = core_properties.revision
            docx_meta["DOCX: Created"] = core_properties.created
            docx_meta["DOCX: Modified"] = core_properties.modified
            docx_meta["DOCX: Last Printed"] = core_properties.last_printed
            docx_meta["DOCX: Category"] = core_properties.category
            docx_meta["DOCX: Comments"] = core_properties.comments
            docx_meta["DOCX: Content Status"] = core_properties.content_status
            docx_meta["DOCX: Identifier"] = core_properties.identifier
            docx_meta["DOCX: Keywords"] = core_properties.keywords
            docx_meta["DOCX: Language"] = core_properties.language
            docx_meta["DOCX: Subject"] = core_properties.subject
            docx_meta["DOCX: Version"] = core_properties.version

            # Add document statistics if available (often not directly in core_properties)
            # You might need to parse the document content for more stats like word count, etc.
            # For simplicity, we stick to core properties here.

        except Exception as e:
            docx_meta["Error (DOCX Metadata)"] = str(e)
        return docx_meta

    def _get_xlsx_metadata(self, filepath):
        """Extracts metadata from XLSX files using openpyxl."""
        xlsx_meta = {}
        try:
            workbook = openpyxl.load_workbook(filepath)
            properties = workbook.properties

            # Extract common properties
            xlsx_meta["XLSX: Creator"] = properties.creator
            xlsx_meta["XLSX: Last Modified By"] = properties.lastModifiedBy
            xlsx_meta["XLSX: Created"] = properties.created
            xlsx_meta["XLSX: Modified"] = properties.modified
            xlsx_meta["XLSX: Title"] = properties.title
            xlsx_meta["XLSX: Subject"] = properties.subject
            xlsx_meta["XLSX: Description"] = properties.description
            xlsx_meta["XLSX: Keywords"] = properties.keywords
            xlsx_meta["XLSX: Category"] = properties.category
            xlsx_meta["XLSX: Manager"] = properties.manager
            xlsx_meta["XLSX: Company"] = properties.company
            xlsx_meta["XLSX: Version"] = properties.version # If available

            # Add sheet names
            xlsx_meta["XLSX: Sheet Names"] = ", ".join(workbook.sheetnames)
            xlsx_meta["XLSX: Number of Sheets"] = len(workbook.sheetnames)

        except Exception as e:
            xlsx_meta["Error (XLSX Metadata)"] = str(e)
        return xlsx_meta

    def _get_pptx_metadata(self, filepath):
        """Extracts metadata from PPTX files using python-pptx."""
        pptx_meta = {}
        try:
            presentation = Presentation(filepath)
            core_properties = presentation.core_properties

            # Extract common core properties
            pptx_meta["PPTX: Title"] = core_properties.title
            pptx_meta["PPTX: Author"] = core_properties.author
            pptx_meta["PPTX: Last Modified By"] = core_properties.last_modified_by
            pptx_meta["PPTX: Revision"] = core_properties.revision
            pptx_meta["PPTX: Created"] = core_properties.created
            pptx_meta["PPTX: Modified"] = core_properties.modified
            pptx_meta["PPTX: Last Printed"] = core_properties.last_printed
            pptx_meta["PPTX: Category"] = core_properties.category
            pptx_meta["PPTX: Comments"] = core_properties.comments
            pptx_meta["PPTX: Content Status"] = core_properties.content_status
            pptx_meta["PPTX: Identifier"] = core_properties.identifier
            pptx_meta["PPTX: Keywords"] = core_properties.keywords
            pptx_meta["PPTX: Language"] = core_properties.language
            pptx_meta["PPTX: Subject"] = core_properties.subject
            pptx_meta["PPTX: Version"] = core_properties.version

            # Add number of slides
            pptx_meta["PPTX: Number of Slides"] = len(presentation.slides)

        except Exception as e:
            pptx_meta["Error (PPTX Metadata)"] = str(e)
        return pptx_meta

    def _display_metadata(self, metadata_dict):
        """Formats and inserts metadata into the scrolled text widget."""
        for key, value in metadata_dict.items():
            # Handle None values gracefully for display
            display_value = "N/A" if value is None else value
            self.metadata_display.insert(tk.END, f"{key:<30}: {display_value}\n")
        self.metadata_display.insert(tk.END, "\n--- End of Metadata ---")

# Main application entry point
if __name__ == "__main__":
    root = tk.Tk()
    app = FileMetadataExtractor(root)
    root.mainloop()
